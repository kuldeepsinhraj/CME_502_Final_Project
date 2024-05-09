import openpyxl as xl
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import curve_fit
#"3% TiO2-ZrO2 (10mg basis) (194 mg) in MO1 4-18-23.xlsx"

Kinetics_File=input("Please input the name of your excel File.\n")

Wavelength=[]
Measure1=[]
Measure2=[]
Measure3=[]
Temperary_Conc=[]
wb_obj=xl.load_workbook(filename=Kinetics_File + ".xlsx",data_only=True)
for sheet in wb_obj:
    for i in range(1,sheet.max_row):
        for j in range(1,sheet.max_column):
            if i == 29:
                try:
                    Wavelength.append(float(sheet.cell(row=i,column=j).value)) #Add try once error is discovered
                except ValueError:
                    Wavelength.append(0)
            if i == 30:
                try:
                    Measure1.append(float(sheet.cell(row=i,column=j).value))
                except ValueError:
                    Measure1.append(0)
            if i == 31:
                try:
                    Measure2.append(float(sheet.cell(row=i,column=j).value))
                except ValueError:
                    Measure2.append(0)
            if i == 32:
                try:
                    Measure3.append(float(sheet.cell(row=i,column=j).value))
                except ValueError:
                    Measure3.append(0)
     
# To follow up from the comments made in class about hardcoding this, the peak is not always found at 370 nm throughout the experiment and this is why the data needs to be hard coded in 
    for k in range(0,len(Wavelength)):
        if Wavelength[k]==370:
            break
    for g in range(0,len(Wavelength)):
        if Wavelength[g]==750:
            break
    avg=[]
    avg.append((Measure1[k]-Measure1[g]+Measure2[k]-Measure2[g]+Measure3[k]-Measure3[g])/3)
    Temperary_Conc.append((avg[-1]-.04)/5751)
    Wavelength.clear()
    Measure1.clear()
    Measure2.clear()
    Measure3.clear()

Conc=[]
for i in range(1,len(Temperary_Conc)):
    Conc.append(Temperary_Conc[-i])
Conc.append(Temperary_Conc[0])
Conc.pop(0)
CA0=Conc[0] 
Conc_Kin=Conc.copy()
Conc_0th=Conc_Kin.copy()
Conc_1st=[]
Conc_2nd=[]

 
for h in range(0,len(Conc_Kin)):
    Conc_1st.append(np.log(Conc_Kin[h]))
for h in range(0,len(Conc_Kin)):
    Conc_2nd.append(1/Conc_Kin[h])

def Conc_0th_ord(t,k_app):
    y=CA0-k_app*t
    return y
def Conc_1st_ord(t,k_app):
    y=np.log(CA0)-k_app*t
    return y
def Conc_2nd_ord(t,k_app):
    y=(1/CA0)+k_app*t
    return y

Conc_0th_model=[]
Conc_1st_model=[]
Conc_2nd_model=[]



time_0=[0,10,20,30,40,50,60,70,80,90,100,110,120]
time_1=[0,10,20,30,40,50,60,70,80,90,100,110,120]
time_2=[0,10,20,30,40,50,60,70,80,90,100,110,120]

param0,covar0=curve_fit(Conc_0th_ord,time_0,Conc_0th)
param1,covar1=curve_fit(Conc_1st_ord,time_1,Conc_1st)
param2,covar2=curve_fit(Conc_2nd_ord,time_2,Conc_2nd)

for i in range(0,len(time_0)):
    Conc_0th_model.append(Conc_0th_ord(time_0[i],param0[0]))
for i in range(0,len(time_1)):
    Conc_1st_model.append(Conc_1st_ord(time_1[i],param1[0]))
for i in range(0,len(time_2)):
    Conc_2nd_model.append(Conc_2nd_ord(time_2[i],param2[0]))
    
Conc_0th_mean=np.array(Conc_0th).mean()
Conc_1st_mean=np.array(Conc_1st).mean()
Conc_2nd_mean=np.array(Conc_2nd).mean()
data_point_number=[]
residual_0Conc=[]
residual_1Conc=[]
residual_2Conc=[]
SSE0=0
SSE1=0
SSE2=0
SST0=0
SST1=0
SST2=0
for i in range(0,len(Conc_0th)):
    residual_0Conc.append(Conc_0th[i]-Conc_0th_model[i])
    data_point_number.append(i+1)
    SSE0+=((Conc_0th[i]-Conc_0th_model[i])**2)
    SST0+=((Conc_0th[i]-Conc_0th_mean)**2)
    Rsquared0=1-(SSE0/SST0)
for i in range(0,len(Conc_1st)):
    residual_1Conc.append(Conc_1st[i]-Conc_1st_model[i])
    SSE1+=((Conc_1st[i]-Conc_1st_model[i])**2)
    SST1+=((Conc_1st[i]-Conc_1st_mean)**2)
    Rsquared1=1-(SSE1/SST1)
for i in range(0,len(Conc_2nd)):
    residual_2Conc.append(Conc_2nd[i]-Conc_2nd_model[i])
    SSE2+=((Conc_2nd[i]-Conc_2nd_model[i])**2)
    SST2+=((Conc_2nd[i]-Conc_2nd_mean)**2)
    Rsquared2=1-(SSE2/SST2)
    
fig=plt.figure(figsize=(6,12),dpi=100)
plt.subplot(3,1,1)
plt.plot(time_0,Conc_0th,"ro")
plt.plot(time_0,Conc_0th_model,"b-")
plt.xlabel("Time (min)")
plt.ylabel("Concentration")
plt.title("Apparent Rate Constant Determination Plot (0th Order)")
plt.legend(["Original Data","Model"])

plt.subplot(3,1,2)
plt.plot(time_1,Conc_1st,"ro")
plt.plot(time_1,Conc_1st_model,"b-")
plt.xlabel("Time (min)")
plt.ylabel("Logorithmic Concentration")
plt.title("Apparent Rate Constant Determination Plot (1st Order)")
plt.legend(["Original Data","Model"])

plt.subplot(3,1,3)
plt.plot(time_2,Conc_2nd,"ro")
plt.plot(time_2,Conc_2nd_model,"b-")
plt.xlabel("Time (min)")
plt.ylabel("Inverse Concentration")
plt.title("Apparent Rate Constant Determination Plot (2nd Order)")
plt.legend(["Original Data","Model"])
plt.tight_layout()
plt.show()

answer = 0
z=0
print("The data points are found at these times on these graphs are: \n{}".format(time_0))
selection=input("Is there an outlier? Please answer Yes or No.\n")
while answer == 0:
    if selection == "No":
        answer = 1
    elif selection == "Yes":
        while answer == 0:
            if z==0:
                outlier_time_value=input("Please select the time that the outlier has occurred.\n")
                for i in range(0,len(time_0)):
                    if time_0[i]==float(outlier_time_value):
                        time_0.pop(i)
                        time_1.pop(i)
                        time_2.pop(i)
                        Conc_0th.pop(i)
                        Conc_1st.pop(i)
                        Conc_2nd.pop(i)
                        Conc_0th_model.pop(i)
                        Conc_1st_model.pop(i)
                        Conc_2nd_model.pop(i)
                        break
                z+=1
            if z>0:
                selection1=input("Is there another outlier? Please answer Yes or No.\n")
                if selection1 == "No":
                    answer = 1
                elif selection1 == "Yes":
                    outlier_time_value=input("Please select the time that the outlier has occurred.\n")
                    for i in range(0,len(time_0)):
                        if time_0[i]==float(outlier_time_value):
                            time_0.pop(i)
                            time_1.pop(i)
                            time_2.pop(i)
                            Conc_0th.pop(i)
                            Conc_1st.pop(i)
                            Conc_2nd.pop(i)
                            Conc_0th_model.pop(i)
                            Conc_1st_model.pop(i)
                            Conc_2nd_model.pop(i)
                            data_point_number.pop(-1)
                            break
                    z+=1
    else:
        continue

Conc_0th_model.clear()
Conc_1st_model.clear()
Conc_2nd_model.clear()
param0,covar0=curve_fit(Conc_0th_ord,time_0,Conc_0th)
param1,covar1=curve_fit(Conc_1st_ord,time_1,Conc_1st)
param2,covar2=curve_fit(Conc_2nd_ord,time_2,Conc_2nd)
for i in range(0,len(time_0)):
    Conc_0th_model.append(Conc_0th_ord(time_0[i],param0[0]))
for i in range(0,len(time_1)):
    Conc_1st_model.append(Conc_1st_ord(time_1[i],param1[0]))
for i in range(0,len(time_2)):
    Conc_2nd_model.append(Conc_2nd_ord(time_2[i],param2[0]))
Conc_0th_mean=np.array(Conc_0th).mean()
Conc_1st_mean=np.array(Conc_1st).mean()
Conc_2nd_mean=np.array(Conc_2nd).mean()
data_point_number=[]
residual_0Conc.clear()
residual_1Conc.clear()
residual_2Conc.clear()
SSE0=0
SSE1=0
SSE2=0
SST0=0
SST1=0
SST2=0
residual_0Sum=0
residual_1Sum=0
residual_2Sum=0
for i in range(0,len(Conc_0th)):
    residual_0Conc.append(Conc_0th[i]-Conc_0th_model[i])
    residual_0Sum+=residual_0Conc[i]
    data_point_number.append(i+1)
    SSE0+=((Conc_0th[i]-Conc_0th_model[i])**2)
    SST0+=((Conc_0th[i]-Conc_0th_mean)**2)
    Rsquared0=1-(SSE0/SST0)
    variance0=SSE0/(len(Conc_0th)-1)
for i in range(0,len(Conc_1st)):
    residual_1Conc.append(Conc_1st[i]-Conc_1st_model[i])
    residual_1Sum+=residual_1Conc[i]
    SSE1+=((Conc_1st[i]-Conc_1st_model[i])**2)
    SST1+=((Conc_1st[i]-Conc_1st_mean)**2)
    Rsquared1=1-(SSE1/SST1)
    variance1=SSE1/(len(Conc_1st)-1)
for i in range(0,len(Conc_2nd)):
    residual_2Conc.append(Conc_2nd[i]-Conc_2nd_model[i])
    residual_2Sum+=residual_2Conc[i]
    SSE2+=((Conc_2nd[i]-Conc_2nd_model[i])**2)
    SST2+=((Conc_2nd[i]-Conc_2nd_mean)**2)
    Rsquared2=1-(SSE2/SST2)
    variance2=SSE2/(len(Conc_2nd)-1)
    
k_app0=param0[0]
k_app1=param1[0]
k_app2=param2[0]

fig=plt.figure(figsize=(6,12),dpi=100)
plt.subplot(3,1,1)
plt.plot(time_0,Conc_0th,"ro")
plt.plot(time_0,Conc_0th_model,"b-")
plt.xlabel("Time (min)")
plt.ylabel("Concentration")
plt.title("Apparent Rate Constant Determination Plot (0th Order)")
plt.legend(["Original Data","Model"])

plt.subplot(3,1,2)
plt.plot(time_1,Conc_1st,"ro")
plt.plot(time_1,Conc_1st_model,"b-")
plt.xlabel("Time (min)")
plt.ylabel("Logorithmic Concentration")
plt.title("Apparent Rate Constant Determination Plot (1st Order)")
plt.legend(["Original Data","Model"])

plt.subplot(3,1,3)
plt.plot(time_2,Conc_2nd,"ro")
plt.plot(time_2,Conc_2nd_model,"b-")
plt.xlabel("Time (min)")
plt.ylabel("Inverse Concentration")
plt.title("Apparent Rate Constant Determination Plot (2nd Order)")
plt.legend(["Original Data","Model"])
plt.tight_layout()
plt.show()

print("The corresponding {:.3f}, {:.3f}, and {:.3f} for the 0th Order, 1st Order and 2nd Order Plots.".format(Rsquared0,Rsquared1,Rsquared2))

print("Please select 0, 1, or 2 for 0th, 1st and 2nd Order, respectively.")
Order_Selection=input("Which order have you selected?\n")
if float(Order_Selection)==0:
    fig=plt.figure(figsize=(6,12),dpi=100)
    plt.subplot(2,1,1)
    plt.plot(time_0,Conc_0th,"ro")
    plt.plot(time_0,Conc_0th_model,"b-")
    plt.xlabel("Time (min)")
    plt.ylabel("Concentration (mol/L)")
    plt.title("Apparent Rate Constant Determination Plot (0th Order)")
    plt.legend(["Original Data","Model"])
    
    plt.subplot(2,1,2)
    plt.plot(data_point_number,residual_0Conc,"ro")
    plt.ylabel("Residuals")
    plt.title("Residuals Plot")
    print("The apparent rate constant value, R^2, sum of the residuals and variance are \n{:.12f} (mol/(L*min), {:.3f}, {:.12f}, and {}, respectively.".format(k_app0,Rsquared0,residual_0Sum,variance0))

if float(Order_Selection)==1:
    fig=plt.figure(figsize=(6,12),dpi=100)
    plt.subplot(2,1,1)
    plt.plot(time_1,Conc_1st,"ro")
    plt.plot(time_1,Conc_1st_model,"b-")
    plt.xlabel("Time (min)")
    plt.ylabel("logarithmic Concentration")
    plt.title("Apparent Rate Constant Determination Plot (1st Order)")
    plt.legend(["Original Data","Model"])
    
    plt.subplot(2,1,2)
    plt.plot(data_point_number,residual_1Conc,"ro")
    plt.ylabel("Residuals")
    plt.title("Residuals Plot")
    print("The apparent rate constant value, R^2, sum of the residuals and variance are \n{:.7f} 1/(min), {:.3f}, {:.3f}, and {}, respectively.".format(k_app1,Rsquared1,residual_1Sum,variance1))
    
if float(Order_Selection)==2:
    fig=plt.figure(figsize=(6,12),dpi=100)
    plt.subplot(2,1,1)
    plt.plot(time_2,Conc_2nd,"ro")
    plt.plot(time_2,Conc_2nd_model,"b-")
    plt.xlabel("Time (min)")
    plt.ylabel("Inverse Concentration (L/mol)")
    plt.title("Apparent Rate Constant Determination Plot (2nd Order)")
    plt.legend(["Original Data","Model"])
    
    plt.subplot(2,1,2)
    plt.plot(data_point_number,residual_2Conc,"ro")
    plt.ylabel("Residuals")
    plt.title("Residuals Plot")
    print("The apparent rate constant value, R^2, sum of the residuals and variance are \n{:.7f} L/(mol*min), {:.3f}, {:.3f}, and {}, respectively.".format(k_app2,Rsquared2,residual_2Sum,variance2))